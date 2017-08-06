from django.conf.global_settings import MEDIA_ROOT
from openpyxl import load_workbook
import os

dir = os.path.join(MEDIA_ROOT, 'files')
truedir = os.path.abspath(dir)


class excelData(object):
    def __init__(self, book):
        self.book = book
        self.workbook = load_workbook(os.path.join(truedir, self.book))
        # self.workbook = load_workbook(os.path.join('d://code//postman//client//files', self.book))
        self.sheets = self.workbook.get_sheet_names()
        self.sheet = self.workbook.get_sheet_by_name(self.sheets[0])

    def get_url(self, r):
        url = self.sheet.cell(row=r, column=1).value
        return url

    def get_method(self, r):
        method = self.sheet.cell(row=r, column=2).value
        return method

    def get_json(self, r):
        start = "C" + str(r)
        end = "J" + str(r)
        su = []
        s = self.sheet[start:end]
        for t in s:
            for cell in t:
                if cell.value:
                    su.append(cell.value)
        k = su[0::2]
        v = su[1::2]
        dict = {}
        for j in range(len(k)):
            if k != '' and v != '':
                dict[str(k[j])] = str(v[j])
        return dict

    def set_return(self, r, back):
        value = back
        self.sheet.cell(row=r, column=11).value = value
        self.workbook.save(os.path.join(truedir, self.book))

    def set_result(self, r, re):
        value = re
        self.sheet.cell(row=r, column=12).value = value
        self.workbook.save(os.path.join(truedir, self.book))


